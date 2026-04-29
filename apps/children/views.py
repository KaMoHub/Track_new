# apps/children/views.py
from django.http import HttpResponseRedirect, JsonResponse
from django.views import View
from django.shortcuts import render
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.utils import timezone
from .models import Child, Direction, Studio, Teacher, StudioEnrollment, ChildList, TeacherStudioAccess
from django.core.exceptions import ValidationError

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd
from django.views.generic import View
import os

from django.http import HttpResponse
from datetime import datetime, date, timedelta
from django.db.models import Count, Q
from django.utils.timezone import make_aware

import openpyxl


def export_children_excel(request):
    """Экспорт списка детей в Excel с учётом фильтров"""
    # Получаем отфильтрованный queryset (используем ту же логику, что в ChildListView)
    queryset = Child.objects.all()

    # Поиск по отдельным полям
    last_name = request.GET.get('last_name', '').strip()
    first_name = request.GET.get('first_name', '').strip()
    patronymic = request.GET.get('patronymic', '').strip()

    if last_name:
        queryset = queryset.filter(last_name__icontains=last_name)
    if first_name:
        queryset = queryset.filter(first_name__icontains=first_name)
    if patronymic:
        queryset = queryset.filter(patronymic__icontains=patronymic)

    # Дети без студии
    no_studio = request.GET.get('no_studio') == 'on'
    if no_studio:
        queryset = queryset.filter(studioenrollment__isnull=True)

    # Сортировка
    sort_field = request.GET.get('sort', 'last_name')
    if sort_field in ['last_name', 'first_name', 'date_of_birth', 'gender']:
        queryset = queryset.order_by(sort_field)
    elif sort_field in ['-last_name', '-first_name', '-date_of_birth', '-gender']:
        queryset = queryset.order_by(sort_field)
    else:
        queryset = queryset.order_by('last_name', 'first_name')

    # Создаём Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Список детей'

    # Заголовки
    headers = ['Фамилия', 'Имя', 'Отчество', 'Дата рождения', 'Возраст', 'Пол']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Данные
    for row_idx, child in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=child.last_name)
        ws.cell(row=row_idx, column=2, value=child.first_name)
        ws.cell(row=row_idx, column=3, value=child.patronymic or '')
        ws.cell(row=row_idx, column=4, value=child.date_of_birth.strftime('%d.%m.%Y') if child.date_of_birth else '')
        ws.cell(row=row_idx, column=5, value=child.age)
        ws.cell(row=row_idx, column=6, value='Мужской' if child.gender == 'M' else 'Женский')

        # Выравнивание
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

    # Формируем ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="children_list.xlsx"'
    wb.save(response)
    return response

def export_enrollments_excel(request):
    """Экспорт списка детей в студиях в Excel с учётом фильтров"""
    from .models import StudioEnrollment, Teacher, Studio
    from django.db.models import Q
    from datetime import datetime

    # Получаем учебный год из профиля
    if hasattr(request.user, 'profile'):
        start = request.user.profile.academic_year_start
        end = request.user.profile.academic_year_end
        current_academic_year = f"{start}-{end}"
    else:
        from apps.accounts.models import UserProfile
        profile = UserProfile.objects.first()
        if profile:
            current_academic_year = f"{profile.academic_year_start}-{profile.academic_year_end}"
        else:
            current_academic_year = "2025-2026"

    # Базовый queryset
    queryset = StudioEnrollment.objects.select_related(
        'child', 'studio', 'direction', 'teacher'
    ).filter(academic_year=current_academic_year)

    # Фильтр по отчисленным
    show_dismissed = request.GET.get('show_dismissed') == 'on'
    if not show_dismissed:
        queryset = queryset.filter(date_of_dismissal__isnull=True)

    # Учитываем выбранные студии
    studio_ids = request.GET.getlist('studios')
    if studio_ids:
        queryset = queryset.filter(studio_id__in=studio_ids)

    # Поиск
    search = request.GET.get('search', '')
    if search:
        queryset = queryset.filter(
            Q(child__fio__icontains=search) |
            Q(studio__name__icontains=search) |
            Q(direction__name__icontains=search) |
            Q(teacher__user__first_name__icontains=search) |
            Q(teacher__user__last_name__icontains=search) |
            Q(teacher__user__username__icontains=search)
        )

    # Учитываем роль пользователя
    user_role = getattr(request.user, 'role', None)
    if user_role == 'teacher':
        try:
            teacher = Teacher.objects.get(user=request.user)
            queryset = queryset.filter(teacher=teacher)
        except Teacher.DoesNotExist:
            queryset = queryset.none()
    elif user_role not in ['methodist', 'admin']:
        queryset = queryset.none()

    # Сортировка
    sort_by = request.GET.get('sort', 'studio__name')
    sort_order = request.GET.get('order', 'asc')

    sort_field_mapping = {
        'child': 'child__fio',
        'studio': 'studio__name',
        'direction': 'direction__name',
        'teacher': 'teacher__fio'
    }

    if sort_by in sort_field_mapping:
        sort_field = sort_field_mapping[sort_by]
    else:
        sort_field = 'studio__name'

    if sort_order == 'desc':
        queryset = queryset.order_by(f'-{sort_field}')
    else:
        queryset = queryset.order_by(sort_field)

    # Создаём Excel файл
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Дети в студиях'

    # Заголовки
    headers = ['Ребенок', 'Студия', 'Направление', 'Педагог', 'Учебный год', 'Дата записи', 'Дата отчисления']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))

    # Данные
    for row_idx, enrollment in enumerate(queryset, 2):
        ws.cell(row=row_idx, column=1, value=enrollment.child.fio)
        ws.cell(row=row_idx, column=2, value=enrollment.studio.name)
        ws.cell(row=row_idx, column=3, value=enrollment.direction.name)
        ws.cell(row=row_idx, column=4, value=str(enrollment.teacher))
        ws.cell(row=row_idx, column=5, value=enrollment.academic_year)
        ws.cell(row=row_idx, column=6, value=enrollment.enrollment_date.strftime('%d.%m.%Y') if enrollment.enrollment_date else '')
        ws.cell(row=row_idx, column=7, value=enrollment.date_of_dismissal.strftime('%d.%m.%Y') if enrollment.date_of_dismissal else '')

        # Выравнивание
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                 top=Side(style='thin'), bottom=Side(style='thin'))

    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # Формируем ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="studio_children.xlsx"'
    wb.save(response)
    return response

def get_accessible_studios(user):
    """Получение списка доступных студий в зависимости от роли пользователя"""
    if hasattr(user, 'role'):
        if user.role == 'teacher':
            try:
                teacher = Teacher.objects.get(user=user)
                accessible_studio_ids = TeacherStudioAccess.objects.filter(
                    teacher=teacher
                ).values_list('studio_id', flat=True)
                return Studio.objects.filter(id__in=accessible_studio_ids)
            except Teacher.DoesNotExist:
                return Studio.objects.none()
        elif user.role in ['methodist', 'admin']:
            return Studio.objects.all()
    return Studio.objects.none()

def get_accessible_enrollments(user):
    """Получение доступных записей в студиях"""
    accessible_studios = get_accessible_studios(user)
    return StudioEnrollment.objects.filter(studio__in=accessible_studios)

def get_accessible_participations(user):
    """Получение доступных участий"""
    accessible_enrollments = get_accessible_enrollments(user)
    from apps.participation.models import Participation
    return Participation.objects.filter(enrollment__in=accessible_enrollments)

def get_participations_for_period(user, start_date, end_date):
    """Единый источник Participation для всех отчетов."""
    accessible_participations = get_accessible_participations(user)
    return accessible_participations.filter(
        report_date__range=(start_date, end_date)
    ).select_related(
        'child',
        'enrollment__direction',
        'enrollment__studio',
        'enrollment__teacher',
        'event',
        'result_type',
    )


from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.views.generic import View
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import datetime, date
import pandas as pd


class DirectionCompetitionReportView(LoginRequiredMixin, View):
    """Отчет по самым активным детям с достижениями"""

    def get(self, request, *args, **kwargs):

        # Получаем параметры из запроса (месяц, год и топ)
        month = request.GET.get('month', datetime.now().month)
        year = request.GET.get('year', datetime.now().year)
        top_n = int(request.GET.get('top', 20))  # Топ N детей

        try:
            month = int(month)
            year = int(year)
        except (ValueError, TypeError):
            current_date = datetime.now()
            month = current_date.month
            year = current_date.year

        # Определяем даты начала и конца периода
        # Учебный год: с сентября предыдущего года по выбранный месяц текущего года
        if 1 <= month <= 8:  # январь-август
            start_date = date(year - 1, 9, 1)  # сентябрь предыдущего года
        else:  # сентябрь-декабрь
            start_date = date(year, 9, 1)  # сентябрь текущего года

        # Конец периода - конец выбранного месяца
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)

        print(f"Период отчета: {start_date} - {end_date}")


        # Получаем данные
        from apps.participation.models import Participation

        accessible_participations = get_accessible_participations(request.user)

        participations = accessible_participations.filter(
            report_date__range=(start_date, end_date)
        ).select_related(
            'child', 'enrollment__studio', 'enrollment__direction', 'event', 'result_type'
        )

        # Группируем по детям
        children_stats = {}

        for participation in participations:
            child_id = participation.child.id
            if child_id not in children_stats:
                children_stats[child_id] = {
                    'fio': participation.child.fio,
                    'studio': participation.enrollment.studio.name if participation.enrollment.studio else '',
                    'direction': participation.enrollment.direction.name if participation.enrollment.direction else '',
                    'total_participations': 0,
                    'certificates': 0,
                    'diplomas': 0,
                    'unique_events': set(),
                    'levels': set(),
                    'last_participation': participation.report_date
                }

            children_stats[child_id]['total_participations'] += 1
            children_stats[child_id]['unique_events'].add(participation.event_id)

            if participation.event and participation.event.level:
                # Переводим уровень на русский
                russian_level = participation.event.get_level_display()
                children_stats[child_id]['levels'].add(russian_level)

            # Считаем сертификаты и дипломы
            if participation.result_type:
                result_name = participation.result_type.name.lower()
                if 'сертификат' in result_name or 'свидетельство' in result_name:
                    children_stats[child_id]['certificates'] += 1
                elif 'диплом' in result_name:
                    children_stats[child_id]['diplomas'] += 1

        # Сортируем по количеству участий и берем топ N
        top_children = sorted(
            children_stats.values(),
            key=lambda x: x['total_participations'],
            reverse=True
        )[:top_n]

        # Подготавливаем данные для Excel
        rows = []
        for i, child_data in enumerate(top_children, 1):
            row = {
                'Место': i,
                'ФИО ребенка': child_data['fio'],
                'Направление': child_data['direction'],
                'Студия': child_data['studio'],
                'Всего участий': child_data['total_participations'],
                'Сертификаты': child_data['certificates'],
                'Дипломы': child_data['diplomas'],
                'Уровни конкурсов': ', '.join(sorted(child_data['levels'])),  # Сортируем для красоты
                'Последнее участие': child_data['last_participation'].strftime('%d.%m.%Y') if child_data[
                    'last_participation'] else ''
            }
            rows.append(row)

        df = pd.DataFrame(rows)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="top_children_{year}.xlsx"'

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Топ {top_n} детей', index=False)

            # Получаем workbook и worksheet для форматирования
            workbook = writer.book
            worksheet = writer.sheets[f'Топ {top_n} детей']

            # Стили для заголовков
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

            # Стили для данных
            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            data_alignment_center = Alignment(horizontal='center', vertical='center')

            # Устанавливаем ширину колонок
            column_widths = {
                'A': 8,  # Место
                'B': 25,  # ФИО ребенка
                'C': 20,  # Направление
                'D': 20,  # Студия
                'E': 12,  # Всего участий
                'F': 12,  # Сертификаты
                'G': 12,  # Дипломы
                'H': 25,  # Уровни конкурсов
                'I': 15,  # Последнее участие
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Форматируем заголовки
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Форматируем данные
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.font = data_font
                    cell.border = border

                    # Центрируем числовые колонки
                    if cell.column in [1, 5, 6, 7]:  # A, E, F, G - Место, Всего участий, Сертификаты, Дипломы
                        cell.alignment = data_alignment_center
                    else:
                        cell.alignment = data_alignment

            # Устанавливаем высоту строки для заголовков
            worksheet.row_dimensions[1].height = 40

        return response

# Добавляем URL в apps/children/urls.py
# path('reports/competition/', views.DirectionCompetitionReportView.as_view(), name='competition_report'),


class ChildrenStudioUploadView(LoginRequiredMixin, View):
    """Загрузка детей и их распределения по студиям из Excel"""
    template_name = 'children/upload_form.html'

    def get(self, request, *args, **kwargs):
        """Показ формы загрузки файла"""
        print("DEBUG: ChildrenStudioUploadView GET запрос")
        return render(request, self.template_name)

    def post(self, request, *args, **kwargs):
        """Обработка загрузки файла"""
        print("DEBUG: ChildrenStudioUploadView POST запрос")

        # Инициализируем счетчики для отчета
        created_children = 0
        updated_children = 0
        created_enrollments = 0
        skipped_records = 0
        errors = []
        error_details = []  # Для подробного отчета

        # Проверяем, есть ли файл в запросе
        if 'excel_file' not in request.FILES:
            print("DEBUG: Файл excel_file не найден в запросе")
            messages.error(request, 'Пожалуйста, выберите Excel файл для загрузки.')
            return render(request, self.template_name)

        excel_file = request.FILES['excel_file']
        print(f"DEBUG: Загружен файл: {excel_file.name}")

        # Проверяем расширение файла
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            print(f"DEBUG: Неверное расширение файла: {excel_file.name}")
            messages.error(request, 'Пожалуйста, загрузите файл в формате Excel (.xlsx или .xls).')
            return render(request, self.template_name)

        try:
            print("DEBUG: Начинаем чтение Excel файла")
            # Читаем Excel файл
            df = pd.read_excel(excel_file, header=0)  # Читаем с заголовками
            print(f"DEBUG: Файл прочитан. Форма DataFrame: {df.shape}")
            print(f"DEBUG: Колонки в файле: {list(df.columns)}")

            if df.empty:
                print("DEBUG: Файл пустой")
                messages.error(request, 'Excel файл пустой.')
                return render(request, self.template_name)

            # Проверяем обязательные колонки
            required_columns = ['FIO', 'BD', 'Pol']
            missing_columns = [col for col in required_columns if col not in df.columns]
            print(f"DEBUG: Обязательные колонки: {required_columns}")
            print(f"DEBUG: Отсутствующие колонки: {missing_columns}")

            if missing_columns:
                messages.error(
                    request,
                    f'В Excel файле отсутствуют обязательные колонки: {", ".join(missing_columns)}'
                )
                print(f"DEBUG: Ошибка - отсутствуют колонки: {missing_columns}")
                return render(request, self.template_name)

            # Загружаем все записи (без ограничения)
            df_limited = df #.head(10)
            print(f"DEBUG: Будет обработано {len(df_limited)} записей")

            print("DEBUG: Начинаем обработку строк")
            # Обрабатываем строки
            for index, row in df_limited.iterrows():
                print(f"DEBUG: Обработка строки {index + 2}")
                try:
                    # Получаем данные из строки

                    fio = str(row['FIO']).strip() if pd.notna(row['FIO']) else ''
                    bd = row['BD'] if pd.notna(row['BD']) else None
                    pol = str(row['Pol']).strip() if pd.notna(row['Pol']) else 'M'

                    print(f"DEBUG: Данные строки {index + 2}:")
                    print(f"  FIO: '{fio}'")
                    print(f"  BD: {bd}")
                    print(f"  Pol: '{pol}'")


                    # Пропускаем пустые строки
                    if not fio:
                        print(f"DEBUG: Пропущена пустая строка {index + 2}")
                        skipped_records += 1
                        continue

                    # Преобразуем дату рождения
                    date_of_birth = None
                    if bd is not None:
                        print(f"DEBUG: Обработка даты рождения: {bd} (тип: {type(bd)})")

                        if isinstance(bd, pd.Timestamp):
                            # Pandas Timestamp
                            date_of_birth = bd.date()
                            print(f"DEBUG: Преобразована дата рождения из Timestamp: {date_of_birth}")
                        elif isinstance(bd, str):
                            # Строка - пытаемся преобразовать
                            try:
                                # Пробуем разные форматы дат
                                date_formats = ['%d.%m.%Y', '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']
                                for fmt in date_formats:
                                    try:
                                        date_of_birth = datetime.strptime(bd, fmt).date()
                                        print(f"DEBUG: Преобразована дата рождения из строки: {date_of_birth}")
                                        break
                                    except ValueError:
                                        continue
                            except Exception as date_error:
                                print(f"DEBUG: Ошибка преобразования даты рождения: {date_error}")
                                date_of_birth = None
                        elif hasattr(bd, 'date'):
                            # datetime.datetime или другие объекты с методом date()
                            try:
                                date_of_birth = bd.date()
                                print(f"DEBUG: Преобразована дата рождения через метод date(): {date_of_birth}")
                            except Exception as date_error:
                                print(f"DEBUG: Ошибка преобразования даты через метод date(): {date_error}")
                                date_of_birth = None
                        else:
                            print(f"DEBUG: Неизвестный тип даты рождения: {type(bd)}")
                            date_of_birth = None

                    print(f"DEBUG: Итоговая дата рождения: {date_of_birth}")

                    # Преобразуем пол
                    gender = 'Мужской'  # По умолчанию мужской
                    # Преобразуем пол
                    gender = 'Мужской'  # По умолчанию мужской
                    if pol:
                        pol_lower = str(pol).lower().strip()

                        # Мужские варианты
                        male_patterns = ['м', 'муж', 'мужской', 'male', 'm', 'муж.', 'мальчик', 'м.', 'мр']

                        # Женские варианты
                        female_patterns = ['д', 'жен', 'женский', 'female', 'f', 'жен.', 'девочка', 'ж.', 'жс']

                        if pol_lower in male_patterns:
                            gender = 'M'
                        elif pol_lower in female_patterns:
                            gender = 'G'
                        else:
                            # Если значение не распознано, оставляем по умолчанию и логируем
                            print(f"DEBUG: Неизвестное значение пола: '{pol}' (приведено к: '{pol_lower}')")
                            # Можно добавить ошибку в лог, если нужно
                            # errors.append(f'Строка {index + 2}: Неизвестное значение пола: "{pol}"')

                    print(f"DEBUG: Преобразованный пол: {gender} (исходное значение: '{pol}')")

                    # Создаем или обновляем ребенка (БЕЗ ПОЛЯ age)
                    n1, n2, n3 = None, None, None

                    space = fio.split()
                    print(space)
                    n1 = space[0]
                    n2 = space[1]
                    space.pop(0)
                    space.pop(0)
                    n3 = ''.join(space)
                    child, child_created = Child.objects.get_or_create(
                        last_name = n1,
                        first_name = n2,
                        patronymic = n3,
                        fio=fio,
                        date_of_birth=date_of_birth,
                        defaults={
                            'gender': gender,
                        }
                    )

                    # Обновляем данные ребенка, если он уже существует
                    if not child_created:
                        child.gender = gender
                        child.save()
                        updated_children += 1
                        print(f"DEBUG: Обновлен ребенок: {fio}")
                    else:
                        created_children += 1
                        print(f"DEBUG: Создан ребенок: {fio}")

                    # Обрабатываем направление и студию - ищем в базе

                except Exception as e:
                    error_msg = f'Ошибка в строке {index + 2}: {str(e)}'
                    print(f"DEBUG: {error_msg}")
                    import traceback
                    print(f"DEBUG: Traceback: {traceback.format_exc()}")
                    errors.append(error_msg)
                    error_details.append(f'Строка {index + 2}: {error_msg}')

            # Показываем результаты
            print(f"DEBUG: Итоги загрузки:")
            print(f"  Создано детей: {created_children}")
            print(f"  Обновлено детей: {updated_children}")
            print(f"  Пропущено записей: {skipped_records}")
            print(f"  Ошибок: {len(errors)}")

            # Создаем отчет о загрузке
            report_content = self.generate_report(
                created_children, updated_children, created_enrollments,
                skipped_records, errors, error_details, excel_file.name
            )

            # Сохраняем отчет в файл
            report_filename = self.save_report(report_content, excel_file.name)

            if created_children > 0:
                messages.success(
                    request,
                    f'Успешно создано {created_children} детей.'
                )

            if updated_children > 0:
                messages.success(
                    request,
                    f'Успешно обновлено {updated_children} детей.'
                )



            if skipped_records > 0:
                messages.info(
                    request,
                    f'Пропущено {skipped_records} записей (пустые или уже существующие).'
                )

            if errors:
                for error in errors:
                    messages.error(request, error)

            if created_children == 0 and updated_children == 0 and created_enrollments == 0 and not errors:
                messages.info(request, 'Нет данных для загрузки.')

            # Добавляем сообщение о сохранении отчета
            if report_filename:
                messages.info(
                    request,
                    f'Отчет о загрузке сохранен в файл: {report_filename}'
                )

        except Exception as e:
            print(f"DEBUG: Ошибка при чтении файла: {e}")
            import traceback
            print(f"DEBUG: Traceback: {traceback.format_exc()}")
            messages.error(request, f'Ошибка при чтении Excel файла: {str(e)}')
            return render(request, self.template_name)

        return HttpResponseRedirect(reverse_lazy('children:studio_children'))

    def generate_report(self, created_children, updated_children, created_enrollments,
                        skipped_records, errors, error_details, filename):
        """Генерация отчета о загрузке"""
        from datetime import datetime

        report_lines = []
        report_lines.append("=" * 50)
        report_lines.append("ОТЧЕТ О ЗАГРУЗКЕ ДАННЫХ ИЗ EXCEL")
        report_lines.append("=" * 50)
        report_lines.append(f"Файл: {filename}")
        report_lines.append(f"Дата загрузки: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}")
        report_lines.append(f"Пользователь: {self.request.user.username}")
        report_lines.append("")
        report_lines.append("ИТОГОВЫЕ РЕЗУЛЬТАТЫ:")
        report_lines.append("-" * 30)
        report_lines.append(f"Создано детей: {created_children}")
        report_lines.append(f"Обновлено детей: {updated_children}")
        report_lines.append(f"Создано записей в студиях: {created_enrollments}")
        report_lines.append(f"Пропущено записей: {skipped_records}")
        report_lines.append(f"Ошибок: {len(errors)}")
        report_lines.append("")

        if error_details:
            report_lines.append("ПОДРОБНЫЕ ОШИБКИ:")
            report_lines.append("-" * 30)
            for error in error_details:
                report_lines.append(f"  {error}")
            report_lines.append("")

        report_lines.append("=" * 50)
        report_lines.append("КОНЕЦ ОТЧЕТА")
        report_lines.append("=" * 50)

        return "\n".join(report_lines)

    def save_report(self, report_content, filename):
        """Сохранение отчета в файл"""
        try:
            from datetime import datetime
            import os

            # Создаем директорию для отчетов, если она не существует
            reports_dir = os.path.join('media', 'uploads', 'reports')
            os.makedirs(reports_dir, exist_ok=True)

            # Генерируем имя файла для отчета
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            base_filename = os.path.splitext(filename)[0]
            report_filename = f"report_{base_filename}_{timestamp}.txt"
            report_path = os.path.join(reports_dir, report_filename)

            # Сохраняем отчет
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(report_content)

            print(f"DEBUG: Отчет сохранен в файл: {report_path}")
            return report_filename

        except Exception as e:
            print(f"DEBUG: Ошибка при сохранении отчета: {e}")
            return None

class ProtectedDeleteMixin:
    """Миксин для проверки возможности удаления объектов"""

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Проверяем, есть ли связанные объекты
        if hasattr(self.object, 'can_be_deleted'):
            if not self.object.can_be_deleted():
                messages.error(
                    request,
                    self.get_deletion_error_message()
                )
                return HttpResponseRedirect(self.get_success_url())

        return super().post(request, *args, **kwargs)

    def get_deletion_error_message(self):
        """Сообщение об ошибке удаления"""
        return f'Невозможно удалить {self.object}. Существуют связанные записи.'


class ChildListView(LoginRequiredMixin, ListView):
    """Список детей"""
    model = Child
    template_name = 'children/child_list.html'
    context_object_name = 'children'
    paginate_by = 50

    def get_queryset(self):
        queryset = super().get_queryset()

        # Поиск по отдельным полям
        last_name = self.request.GET.get('last_name', '').strip()
        first_name = self.request.GET.get('first_name', '').strip()
        patronymic = self.request.GET.get('patronymic', '').strip()

        if last_name:
            queryset = queryset.filter(last_name__icontains=last_name)
        if first_name:
            queryset = queryset.filter(first_name__icontains=first_name)
        if patronymic:
            queryset = queryset.filter(patronymic__icontains=patronymic)

        # Дети без студии
        no_studio = self.request.GET.get('no_studio') == 'on'
        if no_studio:
            queryset = queryset.filter(studioenrollment__isnull=True)

        # Сортировка
        sort_field = self.request.GET.get('sort', 'last_name')
        if sort_field in ['last_name', 'first_name', 'date_of_birth', 'gender']:
            queryset = queryset.order_by(sort_field)
        elif sort_field in ['-last_name', '-first_name', '-date_of_birth', '-gender']:
            queryset = queryset.order_by(sort_field)
        else:
            queryset = queryset.order_by('last_name', 'first_name')


        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_sort'] = self.request.GET.get('sort', 'last_name')
        # Добавляем текущие значения поиска в контекст
        context['search_last_name'] = self.request.GET.get('last_name', '')
        context['search_first_name'] = self.request.GET.get('first_name', '')
        context['search_patronymic'] = self.request.GET.get('patronymic', '')
        return context


# apps/children/views.py (обновляем ChildDetailView)
class ChildDetailView(LoginRequiredMixin, DetailView):
    """Детали ребенка"""
    model = Child
    template_name = 'children/child_detail.html'
    context_object_name = 'child'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем участие ребенка в конкурсах через все его записи в студиях
        from ..participation.models import Participation
        from ..children.models import StudioEnrollment

        # Получаем все записи ребенка в студиях
        enrollments = StudioEnrollment.objects.filter(child=self.object)

        # Получаем все участия через эти записи
        participations = Participation.objects.filter(
            enrollment__in=enrollments
        ).select_related(
            'event', 'result_type', 'enrollment__studio'
        ).order_by('-report_date')

        context['participations'] = participations
        context['participations_count'] = participations.count()
        context['enrollments'] = enrollments

        return context


class ChildCreateView(LoginRequiredMixin, CreateView):
    """Создание ребенка"""
    model = Child
    template_name = 'children/child_form.html'
    # fields = ['fio', 'date_of_birth', 'gender']
    fields = ['last_name', 'first_name', 'patronymic', 'date_of_birth', 'gender']
    success_url = reverse_lazy('children:list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            messages.error(request, 'Доступ только для администраторов')
            return HttpResponseRedirect(reverse_lazy('dashboard:home'))
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Ребенок успешно добавлен.')
        return super().form_valid(form)


class ChildUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование ребенка"""
    model = Child
    template_name = 'children/child_form.html'
    # fields = ['fio', 'date_of_birth', 'gender']
    fields = ['last_name', 'first_name', 'patronymic', 'date_of_birth', 'gender']
    success_url = reverse_lazy('children:list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            messages.error(request, 'Доступ только для администраторов')
            return HttpResponseRedirect(reverse_lazy('dashboard:home'))
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        messages.success(self.request, 'Информация о ребенке успешно обновлена.')
        return super().form_valid(form)


# apps/children/views.py (обновляем ChildDeleteView)
class ChildDeleteView(LoginRequiredMixin, DeleteView):
    """Удаление ребенка"""
    model = Child
    template_name = 'children/child_confirm_delete.html'
    success_url = reverse_lazy('children:list')

    def dispatch(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            messages.error(request, 'Доступ только для администраторов')
            return HttpResponseRedirect(reverse_lazy('dashboard:home'))
        return super().dispatch(request, *args, **kwargs)

    def get_deletion_error_message(self):
        return f'Невозможно удалить ребенка {self.object.fio}, так как он записан в одну или несколько студий. Сначала удалите записи в студиях.'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Проверяем, можно ли удалить ребенка
        if not self.object.can_be_deleted():
            messages.error(
                request,
                self.get_deletion_error_message()
            )
            return HttpResponseRedirect(self.get_success_url())

        messages.success(request, f'Ребенок {self.object.fio} успешно удален.')
        return super().delete(request, *args, **kwargs)

    def get_success_url(self):
        print(f"DEBUG: get_success_url вызван")
        return reverse_lazy('children:list')


# Views для списка детей (справочник)
class ChildListListView(LoginRequiredMixin, ListView):
    model = ChildList
    template_name = 'children/childlist_list.html'
    context_object_name = 'child_lists'
    paginate_by = 50


class ChildListCreateView(LoginRequiredMixin, CreateView):
    model = ChildList
    template_name = 'children/childlist_form.html'
    fields = ['child']
    success_url = reverse_lazy('children:child_list')

    def form_valid(self, form):
        messages.success(self.request, 'Запись в общий список успешно добавлена.')
        return super().form_valid(form)


# Views для направлений
class DirectionListView(LoginRequiredMixin, ListView):
    model = Direction
    template_name = 'children/direction_list.html'
    context_object_name = 'directions'
    paginate_by = 50


class DirectionCreateView(LoginRequiredMixin, CreateView):
    model = Direction
    template_name = 'children/direction_form.html'
    fields = ['name']
    success_url = reverse_lazy('children:direction_list')

    def form_valid(self, form):
        messages.success(self.request, 'Направление успешно добавлено.')
        return super().form_valid(form)



class DirectionDeleteView(LoginRequiredMixin, ProtectedDeleteMixin, DeleteView):
    """Удаление направления"""
    model = Direction
    template_name = 'children/direction_confirm_delete.html'
    success_url = reverse_lazy('children:direction_list')

    def get_deletion_error_message(self):
        return f'Невозможно удалить направление {self.object.name}, так как существуют связанные записи. Сначала удалите связанные данные.'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Проверяем, можно ли удалить направление
        if not self.object.can_be_deleted():
            messages.error(
                request,
                self.get_deletion_error_message()
            )
            return HttpResponseRedirect(reverse_lazy('children:direction_list'))

        messages.success(request, f'Направление {self.object.name} успешно удалено.')
        return super().post(request, *args, **kwargs)




# Views для студий
class StudioListView(LoginRequiredMixin, ListView):
    model = Studio
    template_name = 'children/studio_list.html'
    context_object_name = 'studios'
    paginate_by = 50


class StudioCreateView(LoginRequiredMixin, CreateView):
    model = Studio
    template_name = 'children/studio_form.html'
    fields = ['name', 'direction']
    success_url = reverse_lazy('children:studio_list')

    def form_valid(self, form):
        messages.success(self.request, 'Студия успешно добавлена.')
        return super().form_valid(form)


class StudioDeleteView(LoginRequiredMixin, ProtectedDeleteMixin, DeleteView):
    model = Studio
    template_name = 'children/studio_confirm_delete.html'
    success_url = reverse_lazy('children:studio_list')

    def get_deletion_error_message(self):
        return f'Невозможно удалить студию {self.object.name}, так как в ней есть записи детей.'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Проверяем, можно ли удалить студию
        if not self.object.can_be_deleted():
            messages.error(
                request,
                self.get_deletion_error_message()
            )
            return HttpResponseRedirect(reverse_lazy('children:studio_list'))

        messages.success(request, f'Студия {self.object.name} успешно удалена.')
        return super().post(request, *args, **kwargs)


# Views для педагогов
class TeacherListView(LoginRequiredMixin, ListView):
    model = Teacher
    template_name = 'children/teacher_list.html'
    context_object_name = 'teachers'
    paginate_by = 50


class TeacherCreateView(LoginRequiredMixin, CreateView):
    model = Teacher
    template_name = 'children/teacher_form.html'
    fields = ['user']
    success_url = reverse_lazy('children:teacher_list')

    def form_valid(self, form):
        messages.success(self.request, 'Педагог успешно добавлен.')
        return super().form_valid(form)



class TeacherDeleteView(LoginRequiredMixin, ProtectedDeleteMixin, DeleteView):
    """Удаление педагога"""
    model = Teacher
    template_name = 'children/teacher_confirm_delete.html'
    success_url = reverse_lazy('children:teacher_list')

    def get_deletion_error_message(self):
        return f'Невозможно удалить педагога {self.object}, так как существуют связанные записи. Сначала удалите связанные данные.'

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Проверяем, можно ли удалить педагога
        if not self.object.can_be_deleted():
            messages.error(
                request,
                self.get_deletion_error_message()
            )
            return HttpResponseRedirect(reverse_lazy('children:teacher_list'))

        messages.success(request, f'Педагог {self.object} успешно удален.')
        return super().post(request, *args, **kwargs)




# Views для записей в студии
class EnrollmentListView(LoginRequiredMixin, ListView):
    model = StudioEnrollment
    template_name = 'children/enrollment_list.html'
    context_object_name = 'enrollments'
    paginate_by = 50



# apps/children/views.py (обновляем EnrollmentCreateView)
class EnrollmentCreateView(LoginRequiredMixin, CreateView):
    """Добавление ребенка в студию"""
    model = StudioEnrollment
    template_name = 'children/enrollment_form.html'
    fields = ['child', 'studio', 'teacher', 'academic_year']

    def get_success_url(self):
        return reverse_lazy('children:studio_children')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = self.request.user

        print(f"=== DEBUG: Начало get_form ===")
        print(f"DEBUG: Пользователь: {user.username} (ID: {user.id})")

        # Проверяем роль пользователя
        user_role = getattr(user, 'role', None)
        print(f"DEBUG: Роль пользователя: {user_role}")

        # Всегда показываем всех детей
        all_children = Child.objects.all().order_by('fio')
        form.fields['child'].queryset = all_children
        print(f"DEBUG: Всего детей: {all_children.count()}")

        # Фильтруем студии в зависимости от роли
        if user_role == 'teacher':
            print("DEBUG: Фильтруем студии для педагога")
            # Для педагогов - только доступные студии
            try:
                teacher_record = Teacher.objects.get(user=user)
                accessible_studio_ids = list(TeacherStudioAccess.objects.filter(
                    teacher=teacher_record
                ).values_list('studio_id', flat=True))
                print(f"DEBUG: Доступные студии (ID): {accessible_studio_ids}")

                accessible_studios = Studio.objects.filter(
                    id__in=accessible_studio_ids
                ).order_by('name')
                form.fields['studio'].queryset = accessible_studios
                print(f"DEBUG: Количество доступных студий: {accessible_studios.count()}")
            except Teacher.DoesNotExist:
                print("DEBUG: Пользователь с ролью 'teacher', но не найден в таблице Teacher")
                form.fields['studio'].queryset = Studio.objects.none()
        elif user_role in ['methodist', 'admin']:
            print("DEBUG: Показываем все студии для методиста/админа")
            # Для методистов и админов показываем все студии
            all_studios = Studio.objects.all().order_by('name')
            form.fields['studio'].queryset = all_studios
            print(f"DEBUG: Всего студий: {all_studios.count()}")
        else:
            print("DEBUG: Неизвестная роль, показываем пустой список")
            form.fields['studio'].queryset = Studio.objects.none()

        # Настраиваем поле педагога
        if user_role == 'teacher':
            # Для педагогов пытаемся найти связанного педагога
            try:
                teacher_record = Teacher.objects.get(user=user)
                teacher_queryset = Teacher.objects.filter(id=teacher_record.id)
                form.fields['teacher'].queryset = teacher_queryset
                form.fields['teacher'].initial = teacher_record.id
                print(f"DEBUG: Установлен педагог по умолчанию: {teacher_record.id}")
            except Teacher.DoesNotExist:
                print("DEBUG: Пользователь с ролью 'teacher', но не найден в таблице Teacher")
                form.fields['teacher'].queryset = Teacher.objects.none()
        else:
            # Для методистов и админов показываем всех педагогов
            all_teachers = Teacher.objects.all().order_by('user__last_name')
            form.fields['teacher'].queryset = all_teachers
            print(f"DEBUG: Всего педагогов: {all_teachers.count()}")

        # Устанавливаем текущий учебный год по умолчанию
        academic_year_default = "2024-2025"  # значение по умолчанию
        try:
            if hasattr(user, 'profile'):  # теперь profile, не user_profile
                profile = user.profile
                academic_year_default = f"{profile.academic_year_start}-{profile.academic_year_end}"
                print(f"DEBUG: Учебный год из профиля: {academic_year_default}")
            else:
                print("DEBUG: Используем учебный год по умолчанию")
        except Exception as e:
            print(f"DEBUG: Ошибка при получении учебного года: {e}")

        form.initial['academic_year'] = academic_year_default

        print("=== DEBUG: Конец get_form ===\n")
        return form

    def form_valid(self, form):
        print("=== DEBUG: Начало form_valid ===")

        # Проверяем, не добавлен ли ребенок уже в эту студию в этот год
        child = form.cleaned_data['child']
        studio = form.cleaned_data['studio']
        academic_year = form.cleaned_data['academic_year']

        print(f"DEBUG: Проверка дубликата: ребенок={child.fio}, студия={studio.name}, год={academic_year}")

        if StudioEnrollment.objects.filter(
                child=child,
                studio=studio,
                academic_year=academic_year
        ).exists():
            error_msg = f'Ребенок {child.fio} уже добавлен в студию {studio.name} на {academic_year} учебный год'
            print(f"DEBUG: {error_msg}")
            messages.error(self.request, error_msg)
            return self.form_invalid(form)

        # Автоматически устанавливаем направление на основе студии
        studio = form.cleaned_data['studio']
        form.instance.direction = studio.direction
        print(f"DEBUG: Установлено направление: {studio.direction.name}")

        messages.success(self.request, 'Ребенок успешно добавлен в студию.')
        print("=== DEBUG: Конец form_valid ===")
        return super().form_valid(form)


# Исправляем StudioChildrenListView
# apps/children/views.py (обновляем StudioChildrenListView)
# apps/children/views.py (обновляем StudioChildrenListView)
class StudioChildrenListView(LoginRequiredMixin, ListView):
    """Список детей в студиях с фильтрацией и сортировкой"""
    model = StudioEnrollment
    template_name = 'children/studio_children_list.html'
    context_object_name = 'children'
    paginate_by = 50

    # apps/children/views.py (обновляем get_queryset в StudioChildrenListView)
    def get_queryset(self):
        user = self.request.user
        show_dismissed = self.request.GET.get('show_dismissed', 'off') == 'on'

        # Получаем учебный год из профиля пользователя
        if hasattr(self.request.user, 'profile'):
            start = self.request.user.profile.academic_year_start
            end = self.request.user.profile.academic_year_end
            current_academic_year = f"{start}-{end}"
        else:
            # Если у пользователя нет профиля, берём из любого профиля или значение по умолчанию
            from apps.accounts.models import UserProfile
            profile = UserProfile.objects.first()
            if profile:
                current_academic_year = f"{profile.academic_year_start}-{profile.academic_year_end}"
            else:
                current_academic_year = "2025-2026"

        print(f"=== DEBUG: Начало get_queryset ===")
        print(f"DEBUG: Пользователь: {user.username} (ID: {user.id})")
        queryset = StudioEnrollment.objects.select_related(
            'child', 'studio', 'direction', 'teacher'
        )
        print(f"DEBUG: начальный queryset: {queryset.count()}")

        # Получаем все записи в студиях
        queryset = StudioEnrollment.objects.select_related(
            'child', 'studio',  'direction', 'teacher'
        )

        # Фильтр по учебному году
        print(f"DEBUG: academic_year: {current_academic_year}")
        queryset = queryset.filter(academic_year=current_academic_year)
        print(f"DEBUG: после фильтра по году: {queryset.count()}")

        # Фильтр по отчисленным
        if not show_dismissed:
            queryset = queryset.filter(date_of_dismissal__isnull=True)

        print(f"DEBUG: Всего записей в студиях: {queryset.count()}")

        # Проверяем роль пользователя
        user_role = getattr(user, 'role', None)
        print(f"DEBUG: Роль пользователя: {user_role}")

        # Фильтруем по доступу в зависимости от роли пользователя
        if user_role == 'teacher':
            print("DEBUG: Фильтруем для педагога")
            # Для педагогов - только их студии
            try:
                teacher = Teacher.objects.get(user=user)
                # Получаем записи детей, записанных к этому педагогу
                accessible_enrollments = StudioEnrollment.objects.filter(
                    teacher=teacher
                ).values_list('id', flat=True)
                queryset = queryset.filter(id__in=accessible_enrollments)
            except Teacher.DoesNotExist:
                queryset = queryset.none()
            except Exception:
                queryset = queryset.none()
        elif user_role in ['methodist', 'admin']:
            print("DEBUG: Пользователь - методист или админ, показываем все")
            # Для методистов и админов - все записи
            pass
        else:
            print("DEBUG: Пользователь с неизвестной ролью, показываем пустой список")
            queryset = queryset.none()

        # Фильтры
        studio_ids = self.request.GET.getlist('studios')
        search = self.request.GET.get('search')

        if studio_ids:
            print(f"DEBUG: Фильтр по студии: {studio_ids}")
            queryset = queryset.filter(studio_id__in=studio_ids)
            print(f"DEBUG: Записей после фильтрации по студии: {queryset.count()}")

        if search:
            print(f"DEBUG: Поиск: {search}")
            # ИСПРАВЛЯЕМ поиск - убираем teacher__fio__icontains, так как teacher - ForeignKey
            queryset = queryset.filter(
                Q(child__fio__icontains=search) |
                Q(studio__name__icontains=search) |
                Q(direction__name__icontains=search) |
                Q(teacher__user__first_name__icontains=search) |
                Q(teacher__user__last_name__icontains=search) |
                Q(teacher__user__username__icontains=search)
            )
            print(f"DEBUG: Записей после поиска: {queryset.count()}")

        # Сортировка
        sort_by = self.request.GET.get('sort', 'studio__name')
        sort_order = self.request.GET.get('order', 'asc')

        # Правильные названия полей для сортировки
        sort_field_mapping = {
            'child': 'child__fio',
            'studio': 'studio__name',
            'direction': 'direction__name',
            'teacher': 'teacher__fio'
        }

        # Если переданное поле есть в маппинге, используем его
        if sort_by in sort_field_mapping:
            sort_field = sort_field_mapping[sort_by]
        else:
            sort_field = 'studio__name'  # значение по умолчанию

        # Применяем сортировку
        if sort_order == 'desc':
            queryset = queryset.order_by(f'-{sort_field}')
        else:
            queryset = queryset.order_by(sort_field)

        print(f"DEBUG: Сортировка по {sort_field} {sort_order}")
        print("=== DEBUG: Конец get_queryset ===\n")
        return queryset

    # apps/children/views.py (обновляем get_context_data в StudioChildrenListView)
    # apps/children/views.py (обновляем get_context_data в StudioChildrenListView)
    # apps/children/views.py (исправляем get_context_data в StudioChildrenListView)
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user

        # Получаем текущий отфильтрованный queryset
        filtered_queryset = self.get_queryset()

        # Для фильтров - только доступные данные
        if hasattr(user, 'role'):
            if user.role == 'teacher':
                try:
                    teacher = Teacher.objects.get(user=user)
                    # Только студии этого педагога
                    accessible_studios = TeacherStudioAccess.objects.filter(
                        teacher=teacher
                    ).select_related('studio')
                    studio_ids = [access.studio.id for access in accessible_studios]
                    context['studios'] = Studio.objects.filter(
                        id__in=studio_ids
                    ).order_by('name')
                except (Teacher.DoesNotExist, Exception):
                    context['studios'] = Studio.objects.none()
            elif user.role in ['methodist', 'admin']:
                # Для методистов и админов - все студии
                context['studios'] = Studio.objects.all().order_by('name')
            else:
                context['studios'] = Studio.objects.none()
        else:
            context['studios'] = Studio.objects.none()

        # Передаем текущие значения фильтров
        context['current_studios'] = self.request.GET.getlist('studios')
        context['current_search'] = self.request.GET.get('search', '')

        # Передаем текущие значения сортировки
        context['current_sort'] = self.request.GET.get('sort', 'studio__name')
        context['current_order'] = self.request.GET.get('order', 'asc')

        # Исправленная статистика - учитываем фильтры
        context['total_enrollments'] = filtered_queryset.count()
        context['unique_children_count'] = filtered_queryset.values('child').distinct().count()
        context['studios_count'] = context['studios'].count()

        # Добавляем список разрешенных полей для сортировки
        context['allowed_sort_fields'] = ['child__fio', 'studio__name', 'teacher__fio', 'direction__name']

        # Статистика по отчисленным с учётом роли и выбранных студий
        from django.db.models import Q

        # Получаем учебный год из профиля
        if hasattr(self.request.user, 'profile'):
            start = self.request.user.profile.academic_year_start
            end = self.request.user.profile.academic_year_end
            current_academic_year = f"{start}-{end}"
        else:
            from apps.accounts.models import UserProfile
            profile = UserProfile.objects.first()
            if profile:
                current_academic_year = f"{profile.academic_year_start}-{profile.academic_year_end}"
            else:
                current_academic_year = "2025-2026"

        # Базовый queryset для статистики
        stats_queryset = StudioEnrollment.objects.filter(academic_year=current_academic_year)

        # Учитываем выбранные студии (фильтр)
        selected_studio_ids = self.request.GET.getlist('studios')
        if selected_studio_ids:
            stats_queryset = stats_queryset.filter(studio_id__in=selected_studio_ids)

        # Учитываем роль пользователя
        user_role = getattr(self.request.user, 'role', None)
        if user_role == 'teacher':
            try:
                teacher = Teacher.objects.get(user=self.request.user)
                stats_queryset = stats_queryset.filter(teacher=teacher)
            except Teacher.DoesNotExist:
                stats_queryset = stats_queryset.none()
        elif user_role not in ['methodist', 'admin']:
            stats_queryset = stats_queryset.none()

        # Считаем статистику
        total_enrollments = stats_queryset.count()
        dismissed_count = stats_queryset.filter(date_of_dismissal__isnull=False).count()
        active_count = total_enrollments - dismissed_count

        context['total_enrollments_all'] = total_enrollments
        context['dismissed_count'] = dismissed_count
        context['active_count'] = active_count
        context['current_academic_year'] = current_academic_year

        # # Считаем уникальных детей
        # unique_all = stats_queryset.values('child').distinct().count()
        # unique_active = stats_queryset.filter(date_of_dismissal__isnull=True).values('child').distinct().count()
        # unique_dismissed = stats_queryset.filter(date_of_dismissal__isnull=False).values('child').distinct().count()
        #
        # context['unique_all'] = unique_all
        # context['unique_active'] = unique_active
        # context['unique_dismissed'] = unique_dismissed

        # Считаем уникальных детей
        # Проверяем, есть ли фильтр по студии
        has_studio_filter = bool(selected_studio_ids)

        if has_studio_filter:
            # Если есть фильтр по студии — считаем уникальных по минимальному ID студии
            # Получаем для каждого ребёнка минимальный ID студии среди ВСЕХ его записей (без учёта фильтра)
            all_stats = StudioEnrollment.objects.filter(academic_year=current_academic_year)

            # Учитываем роль пользователя для базового queryset
            if user_role == 'teacher':
                try:
                    teacher = Teacher.objects.get(user=self.request.user)
                    all_stats = all_stats.filter(teacher=teacher)
                except Teacher.DoesNotExist:
                    all_stats = all_stats.none()
            elif user_role not in ['methodist', 'admin']:
                all_stats = all_stats.none()

            # Для каждого ребёнка находим минимальный ID студии
            child_min_studio = {}
            for enrollment in all_stats:
                child_id = enrollment.child_id
                studio_id = enrollment.studio_id
                if child_id not in child_min_studio or studio_id < child_min_studio[child_id]:
                    child_min_studio[child_id] = studio_id

            # Считаем уникальных в отфильтрованной выборке (только те записи, где studio_id == минимальный)
            unique_all = 0
            unique_active = 0
            unique_dismissed = 0

            for enrollment in stats_queryset:
                child_id = enrollment.child_id
                studio_id = enrollment.studio_id
                min_studio = child_min_studio.get(child_id)

                # Учитываем только если эта студия — минимальная для ребёнка
                if min_studio is not None and studio_id == min_studio:
                    unique_all += 1
                    if enrollment.date_of_dismissal:
                        unique_dismissed += 1
                    else:
                        unique_active += 1
        else:
            # Без фильтра по студии — обычный подсчёт уникальных
            unique_all = stats_queryset.values('child').distinct().count()
            unique_active = stats_queryset.filter(date_of_dismissal__isnull=True).values('child').distinct().count()
            unique_dismissed = stats_queryset.filter(date_of_dismissal__isnull=False).values('child').distinct().count()

        context['unique_all'] = unique_all
        context['unique_active'] = unique_active
        context['unique_dismissed'] = unique_dismissed

        return context



# apps/children/views.py (добавляем новые views)

class EnrollmentDetailView(LoginRequiredMixin, DetailView):
    """Просмотр деталей записи в студии"""
    model = StudioEnrollment
    template_name = 'children/enrollment_detail.html'
    context_object_name = 'enrollment'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем участие ребенка в конкурсах
        from ..participation.models import Participation
        participations = Participation.objects.filter(
            enrollment=self.object
        ).select_related(
            'event', 'result_type'
        ).order_by('-report_date')

        context['participations'] = participations
        context['participations_count'] = participations.count()

        return context


# apps/children/views.py (обновляем EnrollmentUpdateView)
class EnrollmentUpdateView(LoginRequiredMixin, UpdateView):
    """Редактирование записи в студии"""
    model = StudioEnrollment
    template_name = 'children/enrollment_form.html'
    fields = ['child', 'studio', 'teacher', 'academic_year', 'date_of_dismissal']
    success_url = reverse_lazy('children:studio_children')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        user = self.request.user

        # Всегда показываем всех детей
        form.fields['child'].queryset = Child.objects.all().order_by('fio')

        # Проверяем роль пользователя
        user_role = getattr(user, 'role', None)

        # Фильтруем студии в зависимости от роли
        if user_role == 'teacher':
            print("DEBUG: Фильтруем студии для педагога при редактировании")
            # Для педагогов - только доступные студии
            try:
                teacher_record = Teacher.objects.get(user=user)
                accessible_studio_ids = list(TeacherStudioAccess.objects.filter(
                    teacher=teacher_record
                ).values_list('studio_id', flat=True))
                print(f"DEBUG: Доступные студии (ID): {accessible_studio_ids}")

                accessible_studios = Studio.objects.filter(
                    id__in=accessible_studio_ids
                ).order_by('name')
                form.fields['studio'].queryset = accessible_studios
                print(f"DEBUG: Количество доступных студий: {accessible_studios.count()}")
            except Teacher.DoesNotExist:
                print("DEBUG: Пользователь с ролью 'teacher', но не найден в таблице Teacher")
                form.fields['studio'].queryset = Studio.objects.none()
        elif user_role in ['methodist', 'admin']:
            print("DEBUG: Показываем все студии для методиста/админа при редактировании")
            # Для методистов и админов показываем все студии
            all_studios = Studio.objects.all().order_by('name')
            form.fields['studio'].queryset = all_studios
            print(f"DEBUG: Всего студий: {all_studios.count()}")
        else:
            print("DEBUG: Неизвестная роль, показываем пустой список")
            form.fields['studio'].queryset = Studio.objects.none()

        # Настраиваем поле педагога
        if user_role == 'teacher':
            # Для педагогов пытаемся найти связанного педагога
            try:
                teacher_record = Teacher.objects.get(user=user)
                teacher_queryset = Teacher.objects.filter(id=teacher_record.id)
                form.fields['teacher'].queryset = teacher_queryset
                # Не устанавливаем initial, чтобы сохранить текущего педагога
                print(f"DEBUG: Установлен педагог по умолчанию: {teacher_record.id}")
            except Teacher.DoesNotExist:
                print("DEBUG: Пользователь с ролью 'teacher', но не найден в таблице Teacher")
                form.fields['teacher'].queryset = Teacher.objects.none()
        else:
            # Для методистов и админов показываем всех педагогов
            all_teachers = Teacher.objects.all().order_by('user__last_name')
            form.fields['teacher'].queryset = all_teachers
            print(f"DEBUG: Всего педагогов: {all_teachers.count()}")

        return form

    def form_valid(self, form):
        # Автоматически устанавливаем направление на основе студии
        studio = form.cleaned_data['studio']
        form.instance.direction = studio.direction

        # Проверяем, не создаст ли это дубликат
        child = form.cleaned_data['child']
        studio = form.cleaned_data['studio']
        academic_year = form.cleaned_data['academic_year']
        current_object = self.get_object()

        if StudioEnrollment.objects.filter(
                child=child,
                studio=studio,
                academic_year=academic_year
        ).exclude(pk=current_object.pk).exists():
            messages.error(
                self.request,
                f'Ребенок {child.fio} уже добавлен в студию {studio.name} на {academic_year} учебный год'
            )
            return self.form_invalid(form)

        messages.success(self.request, 'Запись успешно обновлена.')
        return super().form_valid(form)


# apps/children/views.py (обновляем EnrollmentDeleteView)
class EnrollmentDeleteView(LoginRequiredMixin, DeleteView):
    """Удаление записи из студии"""
    model = StudioEnrollment
    template_name = 'children/enrollment_confirm_delete.html'
    success_url = reverse_lazy('children:studio_children')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()

        # Проверяем, можно ли удалить запись
        if not self.object.can_be_deleted():
            messages.error(
                request,
                f'Невозможно удалить запись ребенка {self.object.child.fio} из студии {self.object.studio.name}, '
                f'так как ребенок участвует в конкурсах. Сначала удалите участия в конкурсах.'
            )
            return HttpResponseRedirect(reverse_lazy('children:enrollment_detail', kwargs={'pk': self.object.pk}))

        child_name = self.object.child.fio
        studio_name = self.object.studio.name

        messages.success(
            request,
            f'Запись ребенка {child_name} из студии {studio_name} успешно удалена.'
        )
        return super().delete(request, *args, **kwargs)


# apps/children/views.py (добавляем новый класс отчета)
class MonthlyAchievementsReportView(LoginRequiredMixin, View):
    """Генерация отчета 'Достижения детей за месяц'"""

    def get(self, request, *args, **kwargs):
        # # Получаем параметры из запроса (месяц и год)
        # month = request.GET.get('month', datetime.now().month)
        # year = request.GET.get('year', datetime.now().year)
        #
        # try:
        #     month = int(month)
        #     year = int(year)
        # except (ValueError, TypeError):
        #     current_date = datetime.now()
        #     month = current_date.month
        #     year = current_date.year
        #
        # # Определяем даты начала и конца месяца
        # start_date = date(year, month, 1)
        # if month == 12:
        #     end_date = date(year + 1, 1, 1) - timedelta(days=1)
        # else:
        #     end_date = date(year, month + 1, 1) - timedelta(days=1)
        #
        # # Название месяца для заголовка
        # month_names = {
        #     1: 'ЯНВАРЬ', 2: 'ФЕВРАЛЬ', 3: 'МАРТ', 4: 'АПРЕЛЬ',
        #     5: 'МАЙ', 6: 'ИЮНЬ', 7: 'ИЮЛЬ', 8: 'АВГУСТ',
        #     9: 'СЕНТЯБРЬ', 10: 'ОКТЯБРЬ', 11: 'НОЯБРЬ', 12: 'ДЕКАБРЬ'
        # }
        # month_name = month_names.get(month, '')

        # 1) Пытаемся взять произвольный период
        start_date_str = request.GET.get('start_date')
        end_date_str = request.GET.get('end_date')

        if start_date_str and end_date_str:
            # формат: YYYY-MM-DD из <input type="date">
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                # если формат кривой – откатываемся к месячному режиму
                start_date = end_date = None
        else:
            start_date = end_date = None

        period_title = ''  # текст для заголовка

        # 2) Если произвольный период не задан – работаем как раньше (месяц/год)
        if not start_date or not end_date:
            month = request.GET.get('month', datetime.now().month)
            year = request.GET.get('year', datetime.now().year)

            try:
                month = int(month)
                year = int(year)
            except (ValueError, TypeError):
                current_date = datetime.now()
                month = current_date.month
                year = current_date.year

            start_date = date(year, month, 1)
            if month == 12:
                end_date = date(year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(year, month + 1, 1) - timedelta(days=1)

            month_names = {
                1: 'ЯНВАРЬ', 2: 'ФЕВРАЛЬ', 3: 'МАРТ', 4: 'АПРЕЛЬ',
                5: 'МАЙ', 6: 'ИЮНЬ', 7: 'ИЮЛЬ', 8: 'АВГУСТ',
                9: 'СЕНТЯБРЬ', 10: 'ОКТЯБРЬ', 11: 'НОЯБРЬ', 12: 'ДЕКАБРЬ'
            }
            month_name = month_names.get(month, '')
            period_title = f'ЗА {month_name} {year}'
            filename_suffix = f'{month:02d}_{year}'
        else:
            # режим произвольного периода
            period_title = f'ЗА ПЕРИОД {start_date.strftime("%d.%m.%Y")} - {end_date.strftime("%d.%m.%Y")}'
            filename_suffix = f'{start_date.strftime("%Y%m%d")}_{end_date.strftime("%Y%m%d")}'

        # Получаем данные участий за указанный месяц с фильтрацией по доступным студиям
        from apps.participation.models import Participation

        # # Получаем доступные участия
        # accessible_participations = get_accessible_participations(request.user)
        #
        # # Фильтруем по периоду
        # participations = accessible_participations.filter(
        #     report_date__range=(start_date, end_date)
        # ).select_related(
        #     'child',
        #     'enrollment__direction',
        #     'enrollment__studio',
        #     'enrollment__teacher',
        #     'event',
        #     'result_type'
        # ).order_by(
        #     'enrollment__direction__name',
        #     'child__fio'
        # )

        participations = get_participations_for_period(
            request.user, start_date, end_date
        ).order_by(
            'enrollment__direction__name',
            'child__fio'
        )

        # Создаем структуру данных для отчета
        report_data = []

        for participation in participations:
            # Получаем все необходимые данные
            child_fio = participation.child.fio
            direction = participation.enrollment.direction.name if participation.enrollment.direction else ''
            studio = participation.enrollment.studio.name if participation.enrollment.studio else ''
            event_name = participation.event.name if participation.event else ''
            event_level = participation.event.get_level_display() if participation.event else ''
            teacher = str(participation.enrollment.teacher) if participation.enrollment.teacher else ''

            # Форма участия
            participation_form = self.get_participation_form(participation)

            # Результат
            result = ''
            if participation.result_type:
                result = participation.result_type.name
            elif participation.custom_result:
                result = participation.custom_result

            # Дата участия
            report_date = participation.report_date.strftime('%d.%m.%Y') if participation.report_date else ''

            report_data.append({
                'child_fio': child_fio,
                'direction': direction,
                'studio': studio,
                'event_name': event_name,
                'event_level': event_level,
                'teacher': teacher,
                'participation_form': participation_form,
                'result': result,
                'report_date': report_date
            })

        # Создаем Excel файл
        df = pd.DataFrame(report_data)

        # Если нет данных, создаем пустой DataFrame с правильными колонками
        if df.empty:
            df = pd.DataFrame(columns=[
                'ФИО ребенка', 'Направление', 'Студия', 'Название мероприятия',
                'Уровень', 'ФИО педагога', 'Форма участия', 'Результат', 'Дата участия'
            ])
        else:
            # Переименовываем колонки на русский
            df = df.rename(columns={
                'child_fio': 'ФИО ребенка',
                'direction': 'Направление',
                'studio': 'Студия',
                'event_name': 'Название мероприятия',
                'event_level': 'Уровень',
                'teacher': 'ФИО педагога',
                'participation_form': 'Форма участия',
                'result': 'Результат',
                'report_date': 'Дата участия'
            })

        # Создаем HttpResponse с Excel файлом
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="dostizheniya_{filename_suffix}.xlsx"'
        )

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            # Основной лист с данными
            df.to_excel(writer, sheet_name='Достижения', index=False)

            # Получаем workbook для форматирования
            workbook = writer.book
            worksheet = writer.sheets['Достижения']

            # Добавляем заголовок
            self.add_report_header(worksheet, period_title, len(report_data))

            # Форматируем таблицу
            self.format_worksheet(worksheet, df)

        return response

    # Остальные методы класса остаются без изменений
    def get_participation_form(self, participation):
        """Определяем форму участия"""
        if participation.event and hasattr(participation.event, 'is_offline'):
            return 'Очная' if participation.event.is_offline else 'Дистанционная'

        event_name = participation.event.name.lower() if participation.event else ''

        if any(word in event_name for word in ['дистанц', 'online', 'онлайн']):
            return 'Дистанционная'
        elif any(word in event_name for word in ['очная', 'offline', 'оффлайн']):
            return 'Очная'
        else:
            return 'Очная'

    def add_report_header(self, worksheet, period_title, records_count):
        """Добавляем заголовок отчета"""
        """Добавляем заголовок отчета (поддерживает месяц или произвольный период)"""
        from openpyxl.styles import Font, Alignment

        worksheet.insert_rows(1, 3)

        worksheet.merge_cells('A1:I1')
        title_cell = worksheet['A1']
        title_cell.value = 'ДОСТИЖЕНИЯ ДЕТЕЙ'
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')

        worksheet.merge_cells('A2:I2')
        period_cell = worksheet['A2']
        period_cell.value = period_title
        period_cell.font = Font(size=12, bold=True)
        period_cell.alignment = Alignment(horizontal='center')

        worksheet.merge_cells('A3:I3')
        count_cell = worksheet['A3']
        count_cell.value = f'Всего записей: {records_count}'
        count_cell.font = Font(size=10)
        count_cell.alignment = Alignment(horizontal='center')

    def format_worksheet(self, worksheet, df):
        """Форматируем внешний вид таблицы"""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        header_font = Font(name='Times New Roman', size=14, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')

        for row_num in range(1, 5):
            for col_num, column_name in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f'{col_letter}{row_num}']
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = header_fill

        column_widths = {
            'A': 30, 'B': 25, 'C': 25, 'D': 50, 'E': 20,
            'F': 40, 'G': 15, 'H': 20, 'I': 12
        }

        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width

        for row_num in range(5, worksheet.max_row + 1):
            for col_num in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col_num)
                cell = worksheet[f'{col_letter}{row_num}']
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(name='Times New Roman', size=12)
                cell.border = thin_border


# apps/children/views.py (обновляем ReportsDashboardView)
# apps/children/views.py (обновляем ReportsDashboardView)
class ReportsDashboardView(LoginRequiredMixin, TemplateView):
    """Дашборд с отчетами"""
    template_name = 'children/reports_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Получаем параметры месяца из запроса или используем текущий
        try:
            month = int(self.request.GET.get('month', 0))
            year = int(self.request.GET.get('year', 0))
        except (ValueError, TypeError):
            month = year = 0

        # Если параметры невалидны, используем текущий месяц
        today = date.today()
        if not (1 <= month <= 12) or not (2020 <= year <= 2100):
            month = today.month
            year = today.year

        # Статистика за выбранный месяц
        start_of_month = date(year, month, 1)
        if month == 12:
            end_of_month = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_of_month = date(year, month + 1, 1) - timedelta(days=1)

        from apps.participation.models import Participation
        monthly_stats = Participation.objects.filter(
            report_date__range=(start_of_month, end_of_month)
        ).aggregate(
            total=Count('id'),
            unique_children=Count('child', distinct=True),
            unique_events=Count('event', distinct=True)
        )

        # Вычисляем текущий квартал
        current_quarter = (month - 1) // 3 + 1

        # Базовые URL для отчетов
        monthly_report_base_url = reverse('children:monthly_achievements_report')
        quarter_report_base_url = reverse('children:semester_achievements_report')

        context.update({
            'current_month': month,
            'current_year': year,
            'current_quarter': current_quarter,
            'monthly_stats': monthly_stats,
            'monthly_report_base_url': monthly_report_base_url,
            'quarter_report_base_url': quarter_report_base_url,
            'reports': [
                {
                    'id': 1,
                    'title': 'Топ активных детей с достижениями',
                    'description': 'Рейтинг самых активных детей по количеству участий в конкурсах',
                    'url': reverse('children:competition_report'),
                    'icon': 'bi-bar-chart',
                    'color': 'primary',
                    'available': True,
                    'needs_params': True,
                    'params_type': 'month_year'
                },
                {
                    'id': 2,
                    'title': 'Достижения детей за месяц',
                    'description': 'Подробный отчет о достижениях детей за выбранный месяц',
                    'url': monthly_report_base_url,
                    'icon': 'bi-trophy',
                    'color': 'success',
                    'available': True,
                    'needs_params': True,
                    'params_type': 'month_year'
                },
                {
                    'id': 3,
                    'title': 'Достижения детей за квартал',
                    'description': 'Количественные показатели по уровням конкурсов с разделением на свидетельства и результаты за квартал',
                    'url': quarter_report_base_url,
                    'icon': 'bi-graph-up',
                    'color': 'info',
                    'available': True,
                    'needs_params': True,
                    'params_type': 'quarter_year'
                }
            ]
        })

        return context


# apps/children/views.py (добавляем новый класс отчета)
class SemesterAchievementsReportView(LoginRequiredMixin, View):
    """Генерация отчета 'Достижения детей за квартал'"""

    def get(self, request, *args, **kwargs):
        # # Получаем параметры из запроса (квартал и год)
        # quarter = request.GET.get('quarter', '1')
        # year = request.GET.get('year', datetime.now().year)
        #
        # try:
        #     quarter = int(quarter)
        #     year = int(year)
        # except (ValueError, TypeError):
        #     current_date = datetime.now()
        #     year = current_date.year
        #     quarter = (current_date.month - 1) // 3 + 1
        #
        # # Определяем даты начала и конца квартала
        # if quarter == 1:
        #     start_date = date(year, 1, 1)
        #     end_date = date(year, 3, 31)
        #     quarter_name = "1 квартал"
        #     months = [1, 2, 3]
        # elif quarter == 2:
        #     start_date = date(year, 4, 1)
        #     end_date = date(year, 6, 30)
        #     quarter_name = "2 квартал"
        #     months = [4, 5, 6]
        # elif quarter == 3:
        #     start_date = date(year, 7, 1)
        #     end_date = date(year, 9, 30)
        #     quarter_name = "3 квартал"
        #     months = [7, 8, 9]
        # else:
        #     start_date = date(year, 10, 1)
        #     end_date = date(year, 12, 31)
        #     quarter_name = "4 квартал"
        #     months = [10, 11, 12]

        # 1) Пытаемся взять произвольный период (как в отчете 2)
        start_date_str = request.GET.get("start_date")
        end_date_str = request.GET.get("end_date")

        start_date = end_date = None
        period_title = ""
        period_is_range = False

        if start_date_str and end_date_str:
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
                period_is_range = True
            except ValueError:
                start_date = end_date = None
                period_is_range = False

        # 2) Если период не задан или некорректен — работаем по кварталу (старое поведение)
        if not start_date or not end_date:
            quarter = request.GET.get("quarter", 1)
            year = request.GET.get("year", datetime.now().year)

            try:
                quarter = int(quarter)
                year = int(year)
            except (ValueError, TypeError):
                current_date = datetime.now()
                year = current_date.year
                quarter = (current_date.month - 1) // 3 + 1

            if quarter == 1:
                start_date = date(year, 1, 1)
                end_date = date(year, 3, 31)
                quarter_name = "1 квартал"
                months = [1, 2, 3]
            elif quarter == 2:
                start_date = date(year, 4, 1)
                end_date = date(year, 6, 30)
                quarter_name = "2 квартал"
                months = [4, 5, 6]
            elif quarter == 3:
                start_date = date(year, 7, 1)
                end_date = date(year, 9, 30)
                quarter_name = "3 квартал"
                months = [7, 8, 9]
            else:
                start_date = date(year, 10, 1)
                end_date = date(year, 12, 31)
                quarter_name = "4 квартал"
                months = [10, 11, 12]

            period_title = f"за {quarter_name} {year} год"
            filename_suffix = f"kvartal_{quarter}_{year}"
        else:
            # режим произвольного диапазона
            # месяцы нужно построить из диапазона (для группировки по месяцам)
            months = sorted(
                {m for m in range(start_date.month, end_date.month + 1)}
            )
            period_title = (
                f"за период {start_date.strftime('%d.%m.%Y')} - "
                f"{end_date.strftime('%d.%m.%Y')}"
            )
            filename_suffix = (
                f"period_{start_date.strftime('%Y%m%d')}_"
                f"{end_date.strftime('%Y%m%d')}"
            )

        # Уровни конкурсов (правильные ключи из модели)
        levels = [
            'center', 'city', 'district', 'republic',
            'regional', 'interregional', 'allrussian', 'international'
        ]

        level_names = {
            'center': 'Центровский',
            'city': 'Городской',
            'district': 'Районный',
            'republic': 'Республиканский',
            'regional': 'Региональный',
            'interregional': 'Межрегиональный',
            'allrussian': 'Всероссийский',
            'international': 'Международный'
        }

        # Названия месяцев для отчета
        month_names = {
            1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
            5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
            9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь'
        }

        # Получаем данные участий за указанный квартал с фильтрацией по доступным студиям
        from apps.participation.models import Participation

        # # Получаем доступные участия
        # accessible_participations = get_accessible_participations(request.user)
        #
        # # Фильтруем по периоду
        # participations = accessible_participations.filter(
        #     report_date__range=(start_date, end_date)
        # ).select_related(
        #     'child',
        #     'enrollment__direction',
        #     'enrollment__studio',
        #     'enrollment__teacher',
        #     'event',
        #     'result_type'
        # ).order_by('child__fio', 'report_date')

        participations = get_participations_for_period(
            request.user, start_date, end_date
        ).order_by('child__fio', 'report_date')

        # Создаем структуру данных для отчета
        report_data = {}

        for participation in participations:
            # Получаем имя педагога
            teacher_name = 'Не указан'
            if participation.enrollment and participation.enrollment.teacher:
                teacher_name = str(participation.enrollment.teacher)

            # Создаем запись для педагога, если ее нет
            if teacher_name not in report_data:
                report_data[teacher_name] = {
                    'teacher': teacher_name,
                    'months': {}
                }

            # Получаем месяц участия
            if participation.report_date:
                month = participation.report_date.month
                month_key = f"{month:02d}"

                # Создаем запись для месяца, если ее нет
                if month_key not in report_data[teacher_name]['months']:
                    report_data[teacher_name]['months'][month_key] = {
                        level: {'s': 0, 'r': 0} for level in levels
                    }

                # Определяем уровень события
                event_level = 'centrovskiy'
                if participation.event and participation.event.level:
                    event_level = participation.event.level

                if event_level not in levels:
                    event_level = 'centrovskiy'

                # Определяем тип результата
                result_type = self.get_result_type(participation)

                # Увеличиваем счетчик
                if result_type == 's':
                    report_data[teacher_name]['months'][month_key][event_level]['s'] += 1
                elif result_type == 'r':
                    report_data[teacher_name]['months'][month_key][event_level]['r'] += 1

        # Создаем правильную структуру данных для DataFrame
        rows = []
        teacher_counter = 1

        for teacher_name, data in report_data.items():
            base_row = {
                '№': teacher_counter,
                'ФИО педагога': teacher_name
            }
            teacher_counter += 1

            # Для каждого месяца и уровня добавляем данные в правильном порядке
            for month_num in months:
                month_key = f"{month_num:02d}"
                month_data = data['months'].get(month_key, {level: {'s': 0, 'r': 0} for level in levels})

                for level in levels:
                    base_row[f'{month_names[month_num]}_{level}_s'] = month_data.get(level, {'s': 0, 'r': 0})['s']
                    base_row[f'{month_names[month_num]}_{level}_r'] = month_data.get(level, {'s': 0, 'r': 0})['r']

            # Добавляем итоги
            base_row['Сертификаты'] = self.calculate_certificates(data, months, levels)
            base_row['Результаты'] = self.calculate_results(data, months, levels)
            base_row['Итого'] = base_row['Сертификаты'] + base_row['Результаты']

            rows.append(base_row)

        # Создаем Excel файл с правильным порядок колонок
        if rows:
            df = pd.DataFrame(rows)
            correct_columns_order = ['№', 'ФИО педагога']

            for level in levels:
                for month_num in months:
                    month_name = month_names[month_num]
                    correct_columns_order.append(f'{month_name}_{level}_s')
                    correct_columns_order.append(f'{month_name}_{level}_r')

            correct_columns_order.extend(['Сертификаты', 'Результаты', 'Итого'])
            df = df[correct_columns_order]
        else:
            columns = ['№', 'ФИО педагога']
            for month_num in months:
                for level in levels:
                    columns.append(f'{month_names[month_num]}_{level}_s')
                    columns.append(f'{month_names[month_num]}_{level}_r')
            columns.extend(['Сертификаты', 'Результаты', 'Итого'])
            df = pd.DataFrame(columns=columns)

        # Создаем HttpResponse с Excel файлом
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = f'attachment; filename="dostizheniya_kvartal_{quarter}_{year}.xlsx"'
        response["Content-Disposition"] = (
            f'attachment; filename="dostizheniya_{filename_suffix}.xlsx"'
        )

        with pd.ExcelWriter(response, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Достижения за квартал', index=False, startrow=6, header=False)
            workbook = writer.book
            worksheet = writer.sheets['Достижения за квартал']

            total_rows = len(df) + 7
            if worksheet.max_row > total_rows:
                for row_num in range(worksheet.max_row, total_rows, -1):
                    worksheet.delete_rows(row_num)

            self.add_report_header(worksheet, period_title, len(rows))
            self.create_complex_header(worksheet, months, month_names, levels, level_names, len(df.columns))
            self.format_quarter_worksheet(worksheet, df, months, month_names, levels, level_names)

        return response

    def get_result_type(self, participation):
        """Определяем тип результата: 's' (свидетельство) или 'r' (результат)"""
        if participation.result_type:
            result_name = participation.result_type.name.lower()
            certificate_keywords = ['сертификат', 'свидетельство', 'участие']

            if any(keyword in result_name for keyword in certificate_keywords):
                return 's'

        if participation.custom_result:
            custom_result = participation.custom_result.lower()
            certificate_keywords = ['сертификат', 'свидетельство', 'участие']
            if any(keyword in custom_result for keyword in certificate_keywords):
                return 's'

        if not participation.result_type and not participation.custom_result:
            return ''

        return 'r'

    def calculate_certificates(self, data, months, levels):
        """Подсчитываем общее количество сертификатов"""
        total = 0
        for month_num in months:
            month_key = f"{month_num:02d}"
            month_data = data['months'].get(month_key, {})
            for level in levels:
                level_data = month_data.get(level, {'s': 0, 'r': 0})
                total += level_data['s']
        return total

    def calculate_results(self, data, months, levels):
        """Подсчитываем общее количество результатов"""
        total = 0
        for month_num in months:
            month_key = f"{month_num:02d}"
            month_data = data['months'].get(month_key, {})
            for level in levels:
                level_data = month_data.get(level, {'s': 0, 'r': 0})
                total += level_data['r']
        return total

    def create_complex_header(self, worksheet, months, month_names, levels, level_names, num_columns):
        """Создает сложную шапку таблицы согласно требованиям"""
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        # Стили
        header_font = Font(name='Times New Roman', size=10, bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='FCD5B4', end_color='FCD5B4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Очищаем возможные старые данные в строках 4-6
        for row in range(4, 7):
            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.value = None
                cell.border = Border()

        # Строка 4: Основные разделы
        worksheet.merge_cells('A4:A6')
        worksheet['A4'].value = '№'
        worksheet.merge_cells('B4:B6')
        worksheet['B4'].value = 'ФИО педагога'

        # Текущая колонка для данных
        current_col = 3

        # Для каждого уровня создаем объединенную ячейку на 6 колонок
        for level in levels:
            level_name = level_names[level]
            start_col = get_column_letter(current_col)
            end_col = get_column_letter(current_col + 5)
            worksheet.merge_cells(f'{start_col}4:{end_col}4')
            cell = worksheet[f'{start_col}4']
            cell.value = level_name
            current_col += 6

        # Итоговые колонки
        start_col = get_column_letter(current_col)
        end_col = get_column_letter(current_col + 2)
        worksheet.merge_cells(f'{start_col}4:{end_col}4')
        worksheet[f'{start_col}4'].value = 'Итог'

        # Строка 5: Месяцы для каждого уровня
        current_col = 3
        for level in levels:
            for month_num in months:
                month_name = month_names[month_num]
                start_col = get_column_letter(current_col)
                end_col = get_column_letter(current_col + 1)
                worksheet.merge_cells(f'{start_col}5:{end_col}5')
                cell = worksheet[f'{start_col}5']
                cell.value = month_name
                current_col += 2

        # Итоговые колонки
        итоговые_заголовки = ['Сертификаты', 'Результаты', 'Итого']
        for i, header in enumerate(итоговые_заголовки):
            col_letter = get_column_letter(current_col + i)
            worksheet.merge_cells(f'{col_letter}5:{col_letter}6')
            worksheet[f'{col_letter}5'].value = header

        # Строка 6: Типы данных
        current_col = 3
        for level in levels:
            for month_num in months:
                col_letter = get_column_letter(current_col)
                worksheet[f'{col_letter}6'].value = 'с'
                current_col += 1

                col_letter = get_column_letter(current_col)
                worksheet[f'{col_letter}6'].value = 'р'
                current_col += 1

        # Применяем стили ко ВСЕМ ячейкам шапки
        for row in range(4, 7):
            for col in range(1, num_columns + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
                if cell.value is not None:
                    cell.fill = header_fill
                else:
                    cell.fill = PatternFill(fill_type=None)

        # Дополнительно: Убедимся, что объединенные ячейки имеют правильные границы
        merged_ranges = [
            'A4:A6', 'B4:B6',
            f'{get_column_letter(3)}4:{get_column_letter(3 + len(levels) * 6 - 1)}4',
            f'{get_column_letter(current_col)}4:{get_column_letter(current_col + 2)}4'
        ]

        # Добавляем диапазоны месяцев
        current_col_temp = 3
        for level in levels:
            for month_num in months:
                start_col = get_column_letter(current_col_temp)
                end_col = get_column_letter(current_col_temp + 1)
                merged_ranges.append(f'{start_col}5:{end_col}5')
                current_col_temp += 2

        # Добавляем итоговые колонки
        for i in range(3):
            col_letter = get_column_letter(current_col + i)
            merged_ranges.append(f'{col_letter}5:{col_letter}6')

        # Применяем стили к объединенным ячейкам
        for range_str in merged_ranges:
            try:
                cell = worksheet[range_str.split(':')[0]]
                cell.border = thin_border
                if cell.value is not None:
                    cell.fill = header_fill
            except:
                pass

    def format_quarter_worksheet(self, worksheet, df, months, month_names, levels, level_names):
        """Форматируем внешний вид таблицы для квартального отчета"""
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Безопасное определение количества колонок
        try:
            if not df.empty:
                num_columns = len(df.columns)
            else:
                num_columns = 2 + (len(levels) * len(months) * 2) + 3
        except:
            num_columns = 50

        # Ограничиваем максимальное количество колонок
        num_columns = min(num_columns, 100)

        # Безопасная настройка ширины колонок
        try:
            worksheet.column_dimensions['A'].width = 5
            worksheet.column_dimensions['B'].width = 30

            for col_num in range(3, num_columns + 1):
                try:
                    col_letter = get_column_letter(col_num)
                    if col_num <= num_columns - 3:
                        worksheet.column_dimensions[col_letter].width = 6
                    else:
                        worksheet.column_dimensions[col_letter].width = 12
                except ValueError:
                    continue
        except:
            pass

        # Безопасное применение стилей к данным
        try:
            start_data_row = 7
            if not df.empty:
                max_rows = min(start_data_row + len(df), worksheet.max_row)
                for row_num in range(start_data_row, max_rows + 1):
                    for col_num in range(1, num_columns + 1):
                        try:
                            cell = worksheet.cell(row=row_num, column=col_num)
                            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            cell.font = Font(name='Times New Roman', size=10)
                            cell.border = thin_border

                            # Специальное выравнивание для колонки ФИО
                            if col_num == 2:
                                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        except:
                            continue
        except:
            pass

    def add_report_header(self, worksheet, period_title, records_count):
        """Добавляем заголовок отчета"""
        from openpyxl.styles import Font, Alignment

        # Очищаем старые данные
        for row in range(1, 4):
            for col in range(1, 50):
                cell = worksheet.cell(row=row, column=col)
                cell.value = None

        worksheet.merge_cells('A1:BA1')
        worksheet['A1'].value = 'Достижения детей'
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')

        worksheet.merge_cells('A2:BA2')
        worksheet['A2'].value = f'за {period_title} год'
        worksheet['A2'].font = Font(size=12, bold=True)
        worksheet['A2'].alignment = Alignment(horizontal='center')

        worksheet.merge_cells('A3:BA3')
        worksheet['A3'].value = f'Всего педагогов: {records_count}'
        worksheet['A3'].font = Font(size=10)
        worksheet['A3'].alignment = Alignment(horizontal='center')


class MonthlyStatsAPIView(LoginRequiredMixin, View):
    """API для получения статистики за выбранный месяц"""

    def get(self, request, *args, **kwargs):
        month = request.GET.get('month')
        year = request.GET.get('year')

        try:
            month = int(month)
            year = int(year)
        except (ValueError, TypeError):
            today = date.today()
            month = today.month
            year = today.year

        # Определяем даты начала и конца месяца
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month + 1, 1) - timedelta(days=1)

        # Получаем статистику участий за указанный месяц с фильтрацией по доступным студиям
        from apps.participation.models import Participation

        # Получаем доступные участия
        accessible_participations = get_accessible_participations(request.user)

        # Фильтруем по периоду
        monthly_stats = accessible_participations.filter(
            report_date__range=(start_date, end_date)
        ).aggregate(
            total=Count('id'),
            unique_children=Count('child', distinct=True),
            unique_events=Count('event', distinct=True)
        )

        # Форматируем название месяца
        from django.utils import timezone
        temp_date = timezone.datetime(year, month, 1)
        month_name = temp_date.strftime('%B')

        return JsonResponse({
            'success': True,
            'month': month,
            'year': year,
            'month_display': f"{month_name} {year}",
            'stats': {
                'total': monthly_stats['total'] or 0,
                'unique_children': monthly_stats['unique_children'] or 0,
                'unique_events': monthly_stats['unique_events'] or 0
            }
        })
