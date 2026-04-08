from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.conf import settings
import sqlite3
import os
import json
from io import BytesIO


@login_required
def admin_dashboard(request):
    """Главная страница админ-инструментов"""
    if request.user.role != 'admin':
        return HttpResponseForbidden("Доступ только для администраторов")

    return render(request, 'admin_tools/dashboard.html')


@login_required
def db_structure(request):
    """Структура БД в текстовом виде"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(f"Файл БД не найден: {db_path}", content_type='text/plain')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    exclude_tables = [
        'accounts_user_groups', 'accounts_user_user_permissions',
        'accounts_useractionlog', 'auth_group', 'auth_group_permissions',
        'auth_permission', 'children_childlist', 'django_migrations',
        'django_content_type', 'django_session', 'django_admin_log',
    ]

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name;")
    tables = cursor.fetchall()

    result_lines = []
    result_lines.append("=" * 100)
    result_lines.append(f"СТРУКТУРА БАЗЫ ДАННЫХ")
    result_lines.append(f"Файл: {os.path.abspath(db_path)}")
    result_lines.append("=" * 100)

    for table in tables:
        table_name = table[0]
        if table_name in exclude_tables:
            continue

        result_lines.append(f"\n{'=' * 100}")
        result_lines.append(f"📋 ТАБЛИЦА: {table_name}")
        result_lines.append(f"{'=' * 100}")

        # Получаем структуру таблицы
        cursor.execute(f"PRAGMA table_info('{table_name}')")
        columns = cursor.fetchall()

        # Получаем внешние ключи
        cursor.execute(f"PRAGMA foreign_key_list('{table_name}')")
        foreign_keys = cursor.fetchall()

        # Создаем словарь связей (колонка -> связанная таблица)
        fk_dict = {}
        for fk in foreign_keys:
            # fk: (id, seq, ref_table, from_col, to_col, on_update, on_delete, match)
            from_col = fk[3]
            ref_table = fk[2]
            fk_dict[from_col] = ref_table

        # Заголовки с новой колонкой
        result_lines.append(f"\n{'Колонка':<35} {'Тип':<15} {'NULL':<8} {'PK':<5} {'Связанная таблица':<30}")
        result_lines.append("-" * 100)

        for col in columns:
            col_id, name, col_type, notnull, default, pk = col
            null_str = "NO" if notnull else "YES"
            pk_str = "✓" if pk else ""
            related_table = fk_dict.get(name, "")  # берем связанную таблицу, если есть

            result_lines.append(f"{name:<35} {col_type:<15} {null_str:<8} {pk_str:<5} {related_table:<30}")

        cursor.execute(f"SELECT COUNT(*) FROM '{table_name}'")
        count = cursor.fetchone()[0]
        result_lines.append(f"\n📊 Всего записей: {count}")

    conn.close()
    return HttpResponse("\n".join(result_lines), content_type='text/plain; charset=utf-8')


@login_required
def project_structure(request):
    """Структура проекта (упрощённая, без рекурсии)"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    root_dir = settings.BASE_DIR

    result_lines = []
    result_lines.append(f"Структура проекта: {root_dir.name}")
    result_lines.append("=" * 50)
    result_lines.append("Для просмотра полной структуры используйте файловый менеджер.")
    result_lines.append(f"Путь к проекту: {root_dir}")

    return HttpResponse("\n".join(result_lines), content_type='text/plain; charset=utf-8')

@login_required
def tables_list(request):
    """Возвращает список всех таблиц (кроме исключенных)"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(json.dumps([]), content_type='application/json')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    exclude_tables = [
        'accounts_user_groups', 'accounts_user_user_permissions',
        'accounts_useractionlog', 'auth_group', 'auth_group_permissions',
        'auth_permission', 'children_childlist', 'django_migrations',
        'django_content_type', 'django_session', 'django_admin_log',
    ]

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name;")
    tables = cursor.fetchall()

    table_list = [t[0] for t in tables if t[0] not in exclude_tables]
    conn.close()

    return HttpResponse(json.dumps(table_list), content_type='application/json')


@login_required
def table_data(request, table_name):
    """Возвращает данные таблицы в формате JSON"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(json.dumps({'error': 'БД не найдена'}), content_type='application/json')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    try:
        # Получаем названия колонок
        cursor.execute(f"PRAGMA table_info('{table_name}')")
        columns_info = cursor.fetchall()
        columns = [col[1] for col in columns_info]

        # Получаем данные
        cursor.execute(f"SELECT * FROM '{table_name}'")
        rows = cursor.fetchall()

        # Преобразуем данные (даты в строки)
        data = []
        for row in rows:
            data_row = []
            for value in row:
                if hasattr(value, 'strftime'):
                    data_row.append(value.strftime('%Y-%m-%d %H:%M:%S'))
                else:
                    data_row.append(str(value) if value is not None else '')
            data.append(data_row)

        conn.close()

        return HttpResponse(json.dumps({
            'columns': columns,
            'rows': data,
            'count': len(rows)
        }), content_type='application/json')

    except Exception as e:
        conn.close()
        return HttpResponse(json.dumps({'error': str(e)}), content_type='application/json')


@login_required
def export_table_excel(request, table_name):
    from openpyxl import Workbook
    """Экспорт таблицы в Excel"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse("БД не найдена", content_type='text/plain')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    try:
        # Получаем названия колонок
        cursor.execute(f"PRAGMA table_info('{table_name}')")
        columns_info = cursor.fetchall()
        columns = [col[1] for col in columns_info]

        # Получаем данные
        cursor.execute(f"SELECT * FROM '{table_name}'")
        rows = cursor.fetchall()

        conn.close()

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = table_name[:31]

        # Заголовки
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Данные
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Настройка ширины колонок
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(col_name)
            for row_idx in range(2, min(len(rows) + 2, 100)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[chr(64 + col_idx)].width = min(max_length + 2, 50)

        # Сохраняем в буфер
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{table_name}.xlsx"'
        return response

    except Exception as e:
        conn.close()
        return HttpResponse(f"Ошибка: {str(e)}", content_type='text/plain')


@login_required
def children_list(request):
    """Возвращает список детей (ID + ФИО)"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(json.dumps([]), content_type='application/json')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute("SELECT id, fio FROM children_child ORDER BY id")
    children = cursor.fetchall()
    conn.close()

    result = []
    for child_id, fio in children:
        result.append({
            'id': child_id,
            'display': f"{child_id} - {fio}" if fio else f"{child_id} - Без имени"
        })

    return HttpResponse(json.dumps(result), content_type='application/json')

@login_required
def search_child_links(request):
    """Ищет все ссылки на выбранного ребенка с детализацией записей"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    child_id = request.GET.get('child_id')

    if not child_id:
        return HttpResponse(json.dumps({'error': 'Не выбран ребенок'}), content_type='application/json')

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(json.dumps({'error': 'БД не найдена'}), content_type='application/json')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Получаем ФИО ребенка
    cursor.execute("SELECT fio, date_of_birth, gender FROM children_child WHERE id = ?", (child_id,))
    child_info = cursor.fetchone()
    if child_info:
        child_fio = child_info[0]
        child_birth = child_info[1]
        child_gender = "Мужской" if child_info[2] == 'M' else "Женский" if child_info[2] == 'F' else "Не указан"
    else:
        child_fio = "Неизвестно"
        child_birth = "Неизвестно"
        child_gender = "Неизвестно"

    exclude_tables = [
        'accounts_user_groups', 'accounts_user_user_permissions',
        'accounts_useractionlog', 'auth_group', 'auth_group_permissions',
        'auth_permission', 'children_childlist', 'django_migrations',
        'django_content_type', 'django_session', 'django_admin_log',
    ]

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name;")
    tables = cursor.fetchall()

    results = []
    total_links = 0

    # Вместо поиска по всем колонкам с 'child' в названии
    # Теперь ищем ТОЛЬКО по полю 'child_id'

    for table in tables:
        table_name = table[0]
        if table_name in exclude_tables:
            continue

        # Получаем структуру таблицы
        cursor.execute(f"PRAGMA table_info('{table_name}')")
        columns = cursor.fetchall()

        # Проверяем, есть ли в таблице колонка 'child_id'
        has_child_id = False
        for col in columns:
            if col[1] == 'child_id':  # col[1] это имя колонки
                has_child_id = True
                break

        if not has_child_id:
            continue

        # Ищем по полю child_id
        try:
            query = f"SELECT * FROM '{table_name}' WHERE child_id = ?"
            cursor.execute(query, (child_id,))
            rows = cursor.fetchall()

            if rows:
                count = len(rows)
                total_links += count

                # Получаем названия всех колонок таблицы
                cursor.execute(f"PRAGMA table_info('{table_name}')")
                all_columns = cursor.fetchall()
                col_names = [c[1] for c in all_columns]

                # Формируем детализацию записей
                records = []
                for row in rows:
                    record = {}
                    for i, value in enumerate(row):
                        if i < len(col_names):
                            if hasattr(value, 'strftime'):
                                value = value.strftime('%Y-%m-%d %H:%M:%S')
                            elif value is None:
                                value = 'NULL'
                            record[col_names[i]] = str(value)
                    records.append(record)

                results.append({
                    'table': table_name,
                    'column': 'child_id',
                    'count': count,
                    'records': records
                })
        except Exception as e:
            pass

    conn.close()

    return HttpResponse(json.dumps({
        'child_id': child_id,
        'child_fio': child_fio,
        'child_birth': str(child_birth) if child_birth else "Не указана",
        'child_gender': child_gender,
        'total_links': total_links,
        'links': results,
        'can_delete': total_links == 0
    }), content_type='application/json')

@login_required
def delete_child(request):
    """Удаляет ребенка (только если нет ссылок)"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    child_id = request.GET.get('child_id')

    if not child_id:
        return HttpResponse(json.dumps({'error': 'Не выбран ребенок'}), content_type='application/json')

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(json.dumps({'error': 'БД не найдена'}), content_type='application/json')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # Получаем ФИО ребенка
    cursor.execute("SELECT fio FROM children_child WHERE id = ?", (child_id,))
    child_info = cursor.fetchone()
    child_fio = child_info[0] if child_info else "Неизвестно"

    # Проверяем наличие ссылок (только по полю child_id)
    exclude_tables = [
        'accounts_user_groups', 'accounts_user_user_permissions',
        'accounts_useractionlog', 'auth_group', 'auth_group_permissions',
        'auth_permission', 'children_childlist', 'django_migrations',
        'django_content_type', 'django_session', 'django_admin_log',
        'children_child'
    ]

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name;")
    tables = cursor.fetchall()

    total_links = 0

    for table in tables:
        table_name = table[0]
        if table_name in exclude_tables:
            continue

        # Проверяем, есть ли в таблице колонка 'child_id'
        cursor.execute(f"PRAGMA table_info('{table_name}')")
        columns = cursor.fetchall()

        has_child_id = False
        for col in columns:
            if col[1] == 'child_id':
                has_child_id = True
                break

        if not has_child_id:
            continue

        try:
            query = f"SELECT COUNT(*) FROM '{table_name}' WHERE child_id = ?"
            cursor.execute(query, (child_id,))
            count = cursor.fetchone()[0]
            total_links += count
        except:
            pass

    if total_links > 0:
        conn.close()
        return HttpResponse(json.dumps({
            'success': False,
            'error': f'Невозможно удалить: найдено {total_links} ссылок'
        }), content_type='application/json')

    # Удаляем ребенка
    cursor.execute("DELETE FROM children_child WHERE id = ?", (child_id,))
    conn.commit()
    conn.close()

    return HttpResponse(json.dumps({
        'success': True,
        'message': f'Ребенок {child_fio} (ID: {child_id}) удален'
    }), content_type='application/json')

@login_required
def replace_links(request):
    """Переносит все ссылки из источника в получатель (пропуская дубли)"""
    if request.user.role != 'admin':
        return HttpResponseForbidden()

    source_id = request.GET.get('source_id')
    target_id = request.GET.get('target_id')

    if not source_id or not target_id:
        return HttpResponse(json.dumps({'error': 'Не выбраны дети'}), content_type='application/json')

    if source_id == target_id:
        return HttpResponse(json.dumps({'error': 'Источник и получатель не могут совпадать'}),
                            content_type='application/json')

    db_path = settings.DATABASES['default']['NAME']

    if not os.path.exists(db_path):
        return HttpResponse(json.dumps({'error': 'БД не найдена'}), content_type='application/json')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    exclude_tables = [
        'accounts_user_groups', 'accounts_user_user_permissions',
        'accounts_useractionlog', 'auth_group', 'auth_group_permissions',
        'auth_permission', 'children_childlist', 'django_migrations',
        'django_content_type', 'django_session', 'django_admin_log',
        'children_child'
    ]

    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name;")
    tables = cursor.fetchall()

    updated_count = 0
    skipped_count = 0
    updated_tables = []
    skipped_details = []

    for table in tables:
        table_name = table[0]
        if table_name in exclude_tables:
            continue

        # Проверяем, есть ли в таблице колонка 'child_id'
        cursor.execute(f"PRAGMA table_info('{table_name}')")
        columns = cursor.fetchall()

        has_child_id = False
        for col in columns:
            if col[1] == 'child_id':
                has_child_id = True
                break

        if not has_child_id:
            continue

        # Получаем все записи, где child_id = source_id
        try:
            cursor.execute(f"SELECT * FROM '{table_name}' WHERE child_id = ?", (source_id,))
            rows = cursor.fetchall()

            if not rows:
                continue

            # Получаем названия колонок
            col_names = [col[1] for col in columns]

            # Определяем, есть ли в таблице unique_together поля (для StudioEnrollment)
            # Это специфичная проверка для таблицы children_studioenrollment
            is_enrollment = (table_name == 'children_studioenrollment')

            for row in rows:
                # Создаем словарь значений
                row_dict = {}
                for i, col_name in enumerate(col_names):
                    row_dict[col_name] = row[i]

                # Проверяем, не существует ли уже такой записи у получателя
                if is_enrollment:
                    # Для children_studioenrollment проверяем уникальность по child, studio, academic_year
                    cursor.execute("""
                        SELECT id FROM children_studioenrollment 
                        WHERE child_id = ? AND studio_id = ? AND academic_year = ?
                    """, (target_id, row_dict.get('studio_id'), row_dict.get('academic_year')))
                    existing = cursor.fetchone()

                    if existing:
                        skipped_count += 1
                        skipped_details.append(
                            f"{table_name}: child_id={source_id} → {target_id} (дубль: studio_id={row_dict.get('studio_id')}, academic_year={row_dict.get('academic_year')})")
                        continue

                # Обновляем child_id на target_id
                try:
                    query = f"UPDATE '{table_name}' SET child_id = ? WHERE child_id = ? AND id = ?"
                    cursor.execute(query, (target_id, source_id, row_dict.get('id')))
                    if cursor.rowcount > 0:
                        updated_count += 1
                except Exception as e:
                    skipped_count += 1
                    skipped_details.append(f"{table_name}: ошибка при обновлении ID={row_dict.get('id')} - {str(e)}")

            if updated_count > 0 or skipped_count > 0:
                updated_tables.append(f"{table_name}: обновлено {updated_count}, пропущено {skipped_count}")

        except Exception as e:
            pass

    conn.commit()
    conn.close()

    return HttpResponse(json.dumps({
        'success': True,
        'updated_count': updated_count,
        'skipped_count': skipped_count,
        'updated_tables': updated_tables,
        'skipped_details': skipped_details
    }), content_type='application/json')
